"""
compare_datasets.py
===================
Compares old preprocessed .npy dataset vs new preprocessed .npy dataset.
Scans all patients in data/patients_preprocessed/ and separates them by
patient ID — old patients are numbered (1-100), new patients are numbered
from 101 onwards (your renamed ISPY-1 patients).

Generates: outputs/dataset_comparison.xlsx
  Sheet 1: per_patient      — one row per patient, all features
  Sheet 2: feature_summary  — statistical comparison old vs new
  Sheet 3: compatibility    — flags any mismatches that need fixing

Run from project root:
    python compare_datasets.py
"""

import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────
PREPROCESSED_ROOT = "data/patients_preprocessed"
OUTPUT_EXCEL      = "outputs/dataset_comparison.xlsx"

# Old patients: IDs 1–100   (numbered folders from original dataset)
# New patients: IDs 101+    (renamed ISPY-1 patients)
OLD_MAX_ID = 100
# ──────────────────────────────────────────────────────────────────

os.makedirs("outputs", exist_ok=True)


def get_patient_group(pid_str):
    """Return 'old_preprocessed' or 'new_ispy1' based on patient ID."""
    try:
        pid_int = int(pid_str)
        return "old_preprocessed" if pid_int <= OLD_MAX_ID else "new_ispy1"
    except ValueError:
        return "new_ispy1"


def analyze_patient(split, pid, pdir):
    """
    Load image.npy and label.npy for one patient and extract all features.
    Returns a dict of metrics.
    """
    img_path = os.path.join(pdir, "image.npy")
    lbl_path = os.path.join(pdir, "label.npy")

    record = {
        "patient_id"    : pid,
        "split"         : split,
        "dataset_group" : get_patient_group(pid),
        "img_path"      : img_path,
    }

    # ── Check files exist ──────────────────────────────────────────
    if not os.path.exists(img_path):
        record["error"] = "image.npy missing"
        return record
    if not os.path.exists(lbl_path):
        record["error"] = "label.npy missing"
        return record

    record["error"] = ""

    # ── Load ──────────────────────────────────────────────────────
    img = np.load(img_path)   # expected (3, D, H, W)
    lbl = np.load(lbl_path)   # expected (1, D, H, W)

    # ── Shape features ────────────────────────────────────────────
    record["img_shape"]    = str(img.shape)
    record["lbl_shape"]    = str(lbl.shape)
    record["n_channels"]   = img.shape[0] if img.ndim == 4 else -1
    record["depth_D"]      = img.shape[1] if img.ndim == 4 else img.shape[0]
    record["height_H"]     = img.shape[2] if img.ndim == 4 else img.shape[1]
    record["width_W"]      = img.shape[3] if img.ndim == 4 else img.shape[2]
    record["total_voxels"] = int(lbl.size)

    # ── Dtype ─────────────────────────────────────────────────────
    record["img_dtype"] = str(img.dtype)
    record["lbl_dtype"] = str(lbl.dtype)

    # ── Image intensity per channel ───────────────────────────────
    for ch in range(min(img.shape[0], 3)):
        ch_data = img[ch]
        record[f"ch{ch}_min"]    = float(np.min(ch_data))
        record[f"ch{ch}_max"]    = float(np.max(ch_data))
        record[f"ch{ch}_mean"]   = float(np.mean(ch_data))
        record[f"ch{ch}_std"]    = float(np.std(ch_data))
        record[f"ch{ch}_p1"]     = float(np.percentile(ch_data, 1))
        record[f"ch{ch}_p99"]    = float(np.percentile(ch_data, 99))
        nz = ch_data[ch_data != 0]
        record[f"ch{ch}_nonzero_mean"] = float(nz.mean()) if len(nz) > 0 else 0.0
        record[f"ch{ch}_nonzero_frac"] = float(len(nz) / ch_data.size)

    # ── Label features ────────────────────────────────────────────
    lbl_sq = lbl.squeeze()
    tumor  = lbl_sq > 0

    record["lbl_unique_vals"]  = str(sorted(np.unique(lbl_sq).tolist()))
    record["lbl_is_binary"]    = bool(set(np.unique(lbl_sq)) <= {0.0, 1.0})
    record["tumor_voxels"]     = int(tumor.sum())
    record["tumor_ratio_pct"]  = float(tumor.sum() / lbl_sq.size * 100)
    record["has_tumor"]        = bool(tumor.sum() > 0)

    # Tumor bounding box
    if tumor.sum() > 0:
        coords = np.argwhere(tumor)
        mins   = coords.min(axis=0)
        maxs   = coords.max(axis=0)
        extent = maxs - mins + 1
        record["tumor_bbox_D"]  = int(extent[0])
        record["tumor_bbox_H"]  = int(extent[1])
        record["tumor_bbox_W"]  = int(extent[2])
        record["tumor_bbox_mm3_approx"] = float(extent[0] * extent[1] * extent[2])
        # Centroid
        record["tumor_centroid_D"] = float(coords.mean(axis=0)[0])
        record["tumor_centroid_H"] = float(coords.mean(axis=0)[1])
        record["tumor_centroid_W"] = float(coords.mean(axis=0)[2])
        # Centroid as fraction of volume (position normalised)
        D, H, W = lbl_sq.shape
        record["tumor_pos_D_norm"] = float(record["tumor_centroid_D"] / D)
        record["tumor_pos_H_norm"] = float(record["tumor_centroid_H"] / H)
        record["tumor_pos_W_norm"] = float(record["tumor_centroid_W"] / W)
    else:
        for k in ["tumor_bbox_D","tumor_bbox_H","tumor_bbox_W",
                  "tumor_bbox_mm3_approx","tumor_centroid_D",
                  "tumor_centroid_H","tumor_centroid_W",
                  "tumor_pos_D_norm","tumor_pos_H_norm","tumor_pos_W_norm"]:
            record[k] = 0.0

    # ── Volume size features ──────────────────────────────────────
    D = record["depth_D"]
    H = record["height_H"]
    W = record["width_W"]
    record["vol_D_x_H"]     = D * H
    record["vol_H_x_W"]     = H * W
    record["aspect_D_to_H"] = float(D / H) if H > 0 else 0.0
    record["aspect_D_to_W"] = float(D / W) if W > 0 else 0.0

    # ── Cross-channel correlation ─────────────────────────────────
    if img.shape[0] >= 2:
        flat0 = img[0].flatten()
        flat1 = img[1].flatten()
        # Sample for speed
        sample_idx = np.random.choice(len(flat0), min(50000, len(flat0)),
                                       replace=False)
        record["corr_ch0_ch1"] = float(
            np.corrcoef(flat0[sample_idx], flat1[sample_idx])[0, 1])
    else:
        record["corr_ch0_ch1"] = 0.0

    if img.shape[0] >= 3:
        flat0 = img[0].flatten()
        flat2 = img[2].flatten()
        sample_idx = np.random.choice(len(flat0), min(50000, len(flat0)),
                                       replace=False)
        record["corr_ch0_ch2"] = float(
            np.corrcoef(flat0[sample_idx], flat2[sample_idx])[0, 1])
        flat1 = img[1].flatten()
        record["corr_ch1_ch2"] = float(
            np.corrcoef(flat1[sample_idx], flat2[sample_idx])[0, 1])
    else:
        record["corr_ch0_ch2"] = 0.0
        record["corr_ch1_ch2"] = 0.0

    return record


def scan_all_patients(root):
    """Walk preprocessed root and return list of (split, pid, pdir)."""
    patients = []
    for split in ["train", "val", "test"]:
        split_dir = os.path.join(root, split)
        if not os.path.exists(split_dir):
            print(f"  [WARN] Split not found: {split_dir}")
            continue
        for pid in sorted(os.listdir(split_dir),
                          key=lambda x: int(x) if x.isdigit() else x):
            pdir = os.path.join(split_dir, pid)
            if os.path.isdir(pdir):
                patients.append((split, pid, pdir))
    return patients


def build_summary(df):
    """Build per-feature statistical summary comparing old vs new."""
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    exclude = ["patient_id"]
    numeric_cols = [c for c in numeric_cols if c not in exclude]

    rows = []
    for col in numeric_cols:
        old_vals = df[df["dataset_group"] == "old_preprocessed"][col].dropna()
        new_vals = df[df["dataset_group"] == "new_ispy1"][col].dropna()

        if len(old_vals) == 0 or len(new_vals) == 0:
            continue

        old_mean = old_vals.mean()
        new_mean = new_vals.mean()
        diff_pct = abs(old_mean - new_mean) / (abs(old_mean) + 1e-9) * 100

        # Compatibility flag
        if diff_pct < 10:
            flag = "COMPATIBLE"
        elif diff_pct < 30:
            flag = "MINOR DIFF"
        elif diff_pct < 70:
            flag = "NOTABLE DIFF"
        else:
            flag = "LARGE DIFF"

        rows.append({
            "feature"       : col,
            "old_mean"      : round(old_mean, 6),
            "old_std"       : round(old_vals.std(), 6),
            "old_min"       : round(old_vals.min(), 6),
            "old_max"       : round(old_vals.max(), 6),
            "new_mean"      : round(new_mean, 6),
            "new_std"       : round(new_vals.std(), 6),
            "new_min"       : round(new_vals.min(), 6),
            "new_max"       : round(new_vals.max(), 6),
            "diff_pct"      : round(diff_pct, 2),
            "compatibility" : flag,
        })

    return pd.DataFrame(rows)


def build_compatibility(df):
    """Generate specific compatibility checks."""
    rows = []

    def check(name, condition, detail, fix=""):
        rows.append({
            "check"    : name,
            "status"   : "PASS" if condition else "FAIL",
            "detail"   : detail,
            "fix_needed": fix,
        })

    old = df[df["dataset_group"] == "old_preprocessed"]
    new = df[df["dataset_group"] == "new_ispy1"]

    # Shape compatibility
    check("n_channels matches",
          (old["n_channels"] == 3).all() and (new["n_channels"] == 3).all(),
          f"Old: {old['n_channels'].unique().tolist()}  "
          f"New: {new['n_channels'].unique().tolist()}",
          "Both must be 3 channels")

    check("Image dtype float32",
          (old["img_dtype"] == "float32").all() and
          (new["img_dtype"] == "float32").all(),
          f"Old dtypes: {old['img_dtype'].unique().tolist()}  "
          f"New dtypes: {new['img_dtype'].unique().tolist()}",
          "Run np.save with .astype(np.float32)")

    check("Label dtype float32",
          (old["lbl_dtype"] == "float32").all() and
          (new["lbl_dtype"] == "float32").all(),
          f"Old: {old['lbl_dtype'].unique().tolist()}  "
          f"New: {new['lbl_dtype'].unique().tolist()}",
          "Run np.save with .astype(np.float32)")

    check("Label is binary (0/1)",
          (old["lbl_is_binary"] == True).all() and
          (new["lbl_is_binary"] == True).all(),
          "Checks unique label values == {0, 1}",
          "Binarize label: (label > 0).astype(np.float32)")

    check("All patients have tumor",
          (old["has_tumor"] == True).all() and
          (new["has_tumor"] == True).all(),
          f"Old missing tumor: {(old['has_tumor']==False).sum()}  "
          f"New missing tumor: {(new['has_tumor']==False).sum()}",
          "Remove patients with 0 tumor voxels")

    # Intensity range
    old_ch1_max = old["ch1_max"].mean()
    new_ch1_max = new["ch1_max"].mean()
    check("Ch1 max < 10 (normalized)",
          old_ch1_max < 10 and new_ch1_max < 10,
          f"Old ch1 max mean: {old_ch1_max:.4f}  "
          f"New ch1 max mean: {new_ch1_max:.4f}",
          "Apply ScaleIntensityRangePercentiles + NormalizeIntensity")

    old_ch1_mean = old["ch1_mean"].mean()
    new_ch1_mean = new["ch1_mean"].mean()
    diff = abs(old_ch1_mean - new_ch1_mean)
    check("Ch1 mean intensity similar (< 0.5 diff)",
          diff < 0.5,
          f"Old ch1 mean: {old_ch1_mean:.4f}  "
          f"New ch1 mean: {new_ch1_mean:.4f}  diff: {diff:.4f}",
          "Re-normalize if diff > 0.5")

    # Volume shapes
    old_D = old["depth_D"].mean()
    new_D = new["depth_D"].mean()
    check("Depth (D) in similar range",
          abs(old_D - new_D) < 80,
          f"Old mean D: {old_D:.1f}  New mean D: {new_D:.1f}",
          "SpatialPadd handles variable sizes automatically")

    # Tumor sizes
    old_tv = old["tumor_voxels"].mean()
    new_tv = new["tumor_voxels"].mean()
    check("Tumor voxel ranges overlap",
          old["tumor_voxels"].max() > new["tumor_voxels"].min(),
          f"Old range: {old['tumor_voxels'].min()}–{old['tumor_voxels'].max()}  "
          f"New range: {new['tumor_voxels'].min()}–{new['tumor_voxels'].max()}",
          "No fix — model handles range via Tversky loss")

    return pd.DataFrame(rows)


def apply_excel_formatting(path):
    """Apply color coding and formatting to the Excel file."""
    wb = load_workbook(path)

    # Color fills
    green_fill  = PatternFill("solid", fgColor="C8E6C9")
    red_fill    = PatternFill("solid", fgColor="FFCDD2")
    amber_fill  = PatternFill("solid", fgColor="FFE0B2")
    yellow_fill = PatternFill("solid", fgColor="FFF9C4")
    blue_fill   = PatternFill("solid", fgColor="BBDEFB")
    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font   = Font(name="Arial", size=9)
    center_aln  = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Header row formatting
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_aln
            cell.border    = border

        # Auto-width columns
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value),
                default=10
            )
            ws.column_dimensions[
                get_column_letter(col[0].column)
            ].width = min(max_len + 4, 35)

        # Body formatting
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font   = body_font
                cell.border = border

        # Sheet-specific color coding
        if sheet_name == "per_patient":
            group_col = None
            for i, cell in enumerate(ws[1], 1):
                if cell.value == "dataset_group":
                    group_col = i
                    break
            if group_col:
                for row in ws.iter_rows(min_row=2):
                    grp = row[group_col - 1].value
                    fill = blue_fill if grp == "old_preprocessed" else green_fill
                    for cell in row:
                        cell.fill = fill

        elif sheet_name == "feature_summary":
            compat_col = None
            for i, cell in enumerate(ws[1], 1):
                if cell.value == "compatibility":
                    compat_col = i
                    break
            if compat_col:
                color_map = {
                    "COMPATIBLE"  : green_fill,
                    "MINOR DIFF"  : yellow_fill,
                    "NOTABLE DIFF": amber_fill,
                    "LARGE DIFF"  : red_fill,
                }
                for row in ws.iter_rows(min_row=2):
                    val = row[compat_col - 1].value
                    if val in color_map:
                        for cell in row:
                            cell.fill = color_map[val]

        elif sheet_name == "compatibility":
            status_col = None
            for i, cell in enumerate(ws[1], 1):
                if cell.value == "status":
                    status_col = i
                    break
            if status_col:
                for row in ws.iter_rows(min_row=2):
                    val = row[status_col - 1].value
                    fill = green_fill if val == "PASS" else red_fill
                    for cell in row:
                        cell.fill = fill

        ws.freeze_panes = "A2"

    wb.save(path)
    print(f"  Formatting applied.")


def main():
    print("=" * 60)
    print("  Dataset Compatibility Comparison")
    print(f"  Root : {PREPROCESSED_ROOT}")
    print("=" * 60)

    # Scan all patients
    all_patients = scan_all_patients(PREPROCESSED_ROOT)
    print(f"\nFound {len(all_patients)} total patient folders")

    old_count = sum(1 for _, pid, _ in all_patients
                    if get_patient_group(pid) == "old_preprocessed")
    new_count = sum(1 for _, pid, _ in all_patients
                    if get_patient_group(pid) == "new_ispy1")
    print(f"  Old patients (ID ≤ {OLD_MAX_ID}) : {old_count}")
    print(f"  New patients (ID > {OLD_MAX_ID}) : {new_count}")

    if len(all_patients) == 0:
        print("\nERROR: No patients found. Check PREPROCESSED_ROOT path.")
        sys.exit(1)

    # Analyze all patients
    print(f"\nAnalyzing {len(all_patients)} patients...")
    records = []
    for split, pid, pdir in tqdm(all_patients, desc="Analyzing"):
        rec = analyze_patient(split, pid, pdir)
        records.append(rec)

    df = pd.DataFrame(records)

    # Report errors
    errors = df[df["error"] != ""]
    if len(errors) > 0:
        print(f"\n  [WARN] {len(errors)} patients had errors:")
        for _, row in errors.iterrows():
            print(f"    Patient {row['patient_id']}: {row['error']}")

    df_clean = df[df["error"] == ""].copy()
    print(f"\n  Successfully analyzed : {len(df_clean)} patients")
    print(f"  Old: {(df_clean['dataset_group']=='old_preprocessed').sum()}")
    print(f"  New: {(df_clean['dataset_group']=='new_ispy1').sum()}")

    # Build summary and compatibility sheets
    print("\nBuilding feature summary...")
    df_summary = build_summary(df_clean)

    print("Running compatibility checks...")
    df_compat = build_compatibility(df_clean)

    # Print quick compatibility result
    passes = (df_compat["status"] == "PASS").sum()
    fails  = (df_compat["status"] == "FAIL").sum()
    print(f"\n  Compatibility checks: {passes} PASS  {fails} FAIL")
    for _, row in df_compat[df_compat["status"] == "FAIL"].iterrows():
        print(f"    FAIL — {row['check']}: {row['detail']}")

    # Write Excel
    print(f"\nWriting Excel report to {OUTPUT_EXCEL}...")
    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
        df_clean.to_excel(writer, sheet_name="per_patient",     index=False)
        df_summary.to_excel(writer, sheet_name="feature_summary", index=False)
        df_compat.to_excel(writer, sheet_name="compatibility",   index=False)

    apply_excel_formatting(OUTPUT_EXCEL)

    print(f"\n{'='*60}")
    print(f"  Report saved : {OUTPUT_EXCEL}")
    print(f"{'='*60}")
    print(f"  Sheets:")
    print(f"    per_patient     — {len(df_clean)} rows, all patient features")
    print(f"    feature_summary — {len(df_summary)} features compared")
    print(f"    compatibility   — {len(df_compat)} checks")
    print(f"\n  Share outputs/dataset_comparison.xlsx")
    print(f"  Once reviewed, run: python combine_and_split.py")


if __name__ == "__main__":
    main()